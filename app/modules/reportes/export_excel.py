"""
app/modules/reportes/export_excel.py

Exportación de reportes a formato Excel usando openpyxl.
INCLUYE COMPARATIVAS SEGÚN TIPO DE REPORTE
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os

from app.modules.comparaciones import model as comparaciones_model
from app.utils.helpers import (
    get_fecha_inicio_semana,
    get_fecha_fin_semana,
    get_fecha_inicio_mes,
    get_fecha_fin_mes
)


def exportar_excel(datos: dict, titulo: str, ruta_archivo: str) -> bool:
    """
    Exporta los datos del reporte a un archivo Excel.
    Incluye comparativas según el tipo de reporte.
    """
    try:
        wb = Workbook()
        
        # ===== HOJA 1: DETALLE DE RETIROS =====
        ws_detalle = wb.active
        ws_detalle.title = "Detalle de Retiros"

        # Estilos
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        total_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        total_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws_detalle.merge_cells('A1:E1')
        ws_detalle['A1'] = titulo
        ws_detalle['A1'].font = Font(size=16, bold=True)
        ws_detalle['A1'].alignment = Alignment(horizontal='center')

        # Período
        ws_detalle.merge_cells('A2:E2')
        ws_detalle['A2'] = f"Período: {datos['fecha_inicio']} al {datos['fecha_fin']}"
        ws_detalle['A2'].alignment = Alignment(horizontal='center')

        # Fecha generación
        ws_detalle.merge_cells('A3:E3')
        ws_detalle['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_detalle['A3'].alignment = Alignment(horizontal='center')

        # Encabezados
        headers = ['ID', 'Fecha', 'Caja', 'Usuario', 'Monto']
        for col, header in enumerate(headers, 1):
            cell = ws_detalle.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row, retiro in enumerate(datos['retiros'], 6):
            ws_detalle.cell(row=row, column=1, value=retiro['id']).border = border
            ws_detalle.cell(row=row, column=2, value=str(retiro['fecha_retiro'])[:16]).border = border
            ws_detalle.cell(row=row, column=3, value=retiro.get('nombre_caja', 'N/A')).border = border
            ws_detalle.cell(row=row, column=4, value=retiro.get('nombre_usuario', retiro.get('username', 'N/A'))).border = border
            monto_cell = ws_detalle.cell(row=row, column=5, value=retiro['monto'])
            monto_cell.number_format = '"$"#,##0.00'
            monto_cell.border = border
            monto_cell.alignment = Alignment(horizontal='right')

        # Total
        total_row = len(datos['retiros']) + 6
        ws_detalle.merge_cells(f'A{total_row}:D{total_row}')
        ws_detalle.cell(row=total_row, column=1, value="TOTAL:").font = Font(bold=True)
        total_cell = ws_detalle.cell(row=total_row, column=5, value=datos['total'])
        total_cell.number_format = '"$"#,##0.00'
        total_cell.font = Font(bold=True)
        total_cell.fill = total_fill
        total_cell.font = total_font
        total_cell.alignment = Alignment(horizontal='right')

        # Ajustar columnas
        column_widths = [8, 18, 20, 20, 15]
        for col, width in enumerate(column_widths, 1):
            ws_detalle.column_dimensions[get_column_letter(col)].width = width

        # ===== DETERMINAR TIPO DE REPORTE =====
        tipo_reporte = ""
        if "Diario" in titulo:
            tipo_reporte = "Diario"
        elif "Semanal" in titulo:
            tipo_reporte = "Semanal"
        elif "Mensual" in titulo:
            tipo_reporte = "Mensual"
        
        # ===== HOJA 2: COMPARATIVAS (SOLO PARA SEMANAL O MENSUAL) =====
        if tipo_reporte in ["Semanal", "Mensual"]:
            ws_comparativas = wb.create_sheet("Comparativas")
            
            # Título
            ws_comparativas.merge_cells('A1:D1')
            ws_comparativas['A1'] = f"Comparativa {tipo_reporte}"
            ws_comparativas['A1'].font = Font(size=14, bold=True)
            ws_comparativas['A1'].alignment = Alignment(horizontal='center')
            
            ws_comparativas['A3'] = "Fecha de generación:"
            ws_comparativas['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M")
            
            row = 5
            
            if tipo_reporte == "Semanal":
                ws_comparativas.merge_cells(f'A{row}:D{row}')
                ws_comparativas.cell(row=row, column=1, value="COMPARATIVA SEMANAL").font = Font(bold=True, size=12, color="1565C0")
                row += 1
                
                try:
                    comparacion = comparaciones_model.comparar_semana_actual_con_anterior()
                    
                    semana_actual_total = comparacion['semana2']['total']
                    semana_anterior_total = comparacion['semana1']['total']
                    diferencia = comparacion['diferencia']
                    porcentaje = comparacion['porcentaje']
                    tendencia = comparacion['tendencia']
                    
                    # Encabezados
                    ws_comparativas.cell(row=row+1, column=1, value="Concepto").font = Font(bold=True)
                    ws_comparativas.cell(row=row+1, column=2, value="Semana Anterior").font = Font(bold=True)
                    ws_comparativas.cell(row=row+1, column=3, value="Semana Actual").font = Font(bold=True)
                    ws_comparativas.cell(row=row+1, column=4, value="Variación").font = Font(bold=True)
                    
                    # Datos
                    ws_comparativas.cell(row=row+2, column=1, value="Total")
                    ws_comparativas.cell(row=row+2, column=2, value=semana_anterior_total).number_format = '"$"#,##0.00'
                    ws_comparativas.cell(row=row+2, column=3, value=semana_actual_total).number_format = '"$"#,##0.00'
                    
                    variacion_texto = f"{'▲' if tendencia == 'aumento' else '▼' if tendencia == 'disminucion' else '='} ${abs(diferencia):,.2f} ({porcentaje:.1f}%)"
                    ws_comparativas.cell(row=row+2, column=4, value=variacion_texto)
                    
                    if tendencia == 'aumento':
                        ws_comparativas.cell(row=row+2, column=4).font = Font(color="2E7D32", bold=True)
                    elif tendencia == 'disminucion':
                        ws_comparativas.cell(row=row+2, column=4).font = Font(color="C62828", bold=True)
                    
                    row += 4
                    
                    # Detalle por caja
                    if 'detalle_cajas' in comparacion and comparacion['detalle_cajas']:
                        ws_comparativas.cell(row=row+1, column=1, value="Detalle por Caja").font = Font(bold=True, italic=True)
                        row += 1
                        
                        ws_comparativas.cell(row=row+1, column=1, value="Caja").font = Font(bold=True)
                        ws_comparativas.cell(row=row+1, column=2, value="Semana Anterior").font = Font(bold=True)
                        ws_comparativas.cell(row=row+1, column=3, value="Semana Actual").font = Font(bold=True)
                        ws_comparativas.cell(row=row+1, column=4, value="Variación").font = Font(bold=True)
                        row += 1
                        
                        for caja in comparacion['detalle_cajas']:
                            ws_comparativas.cell(row=row+1, column=1, value=caja['nombre'])
                            ws_comparativas.cell(row=row+1, column=2, value=caja['total_semana1']).number_format = '"$"#,##0.00'
                            ws_comparativas.cell(row=row+1, column=3, value=caja['total_semana2']).number_format = '"$"#,##0.00'
                            variacion_caja = caja['total_semana2'] - caja['total_semana1']
                            signo = "▲" if variacion_caja > 0 else "▼" if variacion_caja < 0 else "="
                            ws_comparativas.cell(row=row+1, column=4, value=f"{signo} ${abs(variacion_caja):,.2f}")
                            row += 1
                            
                except Exception as e:
                    ws_comparativas.cell(row=row+2, column=1, value="No hay datos suficientes para comparativa semanal")
            
            elif tipo_reporte == "Mensual":
                ws_comparativas.merge_cells(f'A{row}:D{row}')
                ws_comparativas.cell(row=row, column=1, value="COMPARATIVA MENSUAL").font = Font(bold=True, size=12, color="1565C0")
                row += 1
                
                try:
                    comparacion = comparaciones_model.comparar_mes_actual_con_anterior()
                    
                    mes_actual_total = comparacion['mes2']['total']
                    mes_anterior_total = comparacion['mes1']['total']
                    diferencia = comparacion['diferencia']
                    porcentaje = comparacion['porcentaje']
                    tendencia = comparacion['tendencia']
                    
                    # Encabezados
                    ws_comparativas.cell(row=row+1, column=1, value="Concepto").font = Font(bold=True)
                    ws_comparativas.cell(row=row+1, column=2, value="Mes Anterior").font = Font(bold=True)
                    ws_comparativas.cell(row=row+1, column=3, value="Mes Actual").font = Font(bold=True)
                    ws_comparativas.cell(row=row+1, column=4, value="Variación").font = Font(bold=True)
                    
                    # Datos
                    ws_comparativas.cell(row=row+2, column=1, value="Total")
                    ws_comparativas.cell(row=row+2, column=2, value=mes_anterior_total).number_format = '"$"#,##0.00'
                    ws_comparativas.cell(row=row+2, column=3, value=mes_actual_total).number_format = '"$"#,##0.00'
                    
                    variacion_texto = f"{'▲' if tendencia == 'aumento' else '▼' if tendencia == 'disminucion' else '='} ${abs(diferencia):,.2f} ({porcentaje:.1f}%)"
                    ws_comparativas.cell(row=row+2, column=4, value=variacion_texto)
                    
                    if tendencia == 'aumento':
                        ws_comparativas.cell(row=row+2, column=4).font = Font(color="2E7D32", bold=True)
                    elif tendencia == 'disminucion':
                        ws_comparativas.cell(row=row+2, column=4).font = Font(color="C62828", bold=True)
                    
                    row += 4
                    
                    # Detalle por caja
                    if 'detalle_cajas' in comparacion and comparacion['detalle_cajas']:
                        ws_comparativas.cell(row=row+1, column=1, value="Detalle por Caja").font = Font(bold=True, italic=True)
                        row += 1
                        
                        ws_comparativas.cell(row=row+1, column=1, value="Caja").font = Font(bold=True)
                        ws_comparativas.cell(row=row+1, column=2, value="Mes Anterior").font = Font(bold=True)
                        ws_comparativas.cell(row=row+1, column=3, value="Mes Actual").font = Font(bold=True)
                        ws_comparativas.cell(row=row+1, column=4, value="Variación").font = Font(bold=True)
                        row += 1
                        
                        for caja in comparacion['detalle_cajas']:
                            ws_comparativas.cell(row=row+1, column=1, value=caja['nombre'])
                            ws_comparativas.cell(row=row+1, column=2, value=caja['total_mes1']).number_format = '"$"#,##0.00'
                            ws_comparativas.cell(row=row+1, column=3, value=caja['total_mes2']).number_format = '"$"#,##0.00'
                            variacion_caja = caja['total_mes2'] - caja['total_mes1']
                            signo = "▲" if variacion_caja > 0 else "▼" if variacion_caja < 0 else "="
                            ws_comparativas.cell(row=row+1, column=4, value=f"{signo} ${abs(variacion_caja):,.2f}")
                            row += 1
                            
                except Exception as e:
                    ws_comparativas.cell(row=row+2, column=1, value="No hay datos suficientes para comparativa mensual")
            
            # Ajustar columnas en hoja de comparativas
            for col in range(1, 5):
                ws_comparativas.column_dimensions[get_column_letter(col)].width = 20

        # Guardar
        os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)
        wb.save(ruta_archivo)
        return True

    except Exception as e:
        raise Exception(f"Error exportando Excel: {e}")